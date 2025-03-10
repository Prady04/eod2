import yfinance as yf
import pandas as pd
import time
import datetime
import schedule
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill 
import pandas as pd


# Configuration
STOCKS = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'META']  # Default list of stocks
OUTPUT_FILE = 'stock_prices.xlsx'
THRESHOLD_PERCENTAGE = 0.5  # Half a percent threshold

def get_stock_data(stock_symbols):
    """Get real-time stock data for the given list of stock symbols."""
    data = {}
    now = datetime.datetime.now()
    today = now.strftime('%Y-%m-%d')
    
    # Define the 9:30-9:45 time window
    start_time = f"{today} 09:30:00"
    end_time = f"{today} 09:45:00"
    
    df = pd.read_csv('n500.csv')
    for symbol in df['Symbol']:
        try:
            stock = yf.Ticker(symbol+".NS")
            history = stock.history(period='1d', interval='15m')
            
            if not history.empty:
                # Get the day's high and low as of the current time
                morning_candle = history.between_time('09:30', '09:45')
                
                if not morning_candle.empty:
                    # Get the high and low from the 9:30-9:45 time window
                    candle_high = morning_candle['High'].max()
                    candle_low = morning_candle['Low'].min()
                    
                    # Get the current market price (last available price)
                    current_price = history['Close'].iloc[-1]
                    
                    data[symbol] = {
                        'High': candle_high,
                        'Low': candle_low,
                        'CMP': current_price
                    }
                    
                    print(f"{symbol}: 9:30-9:45 High: {candle_high}, Low: {candle_low}, Current: {current_price}")
                else:
                    print(f"No data available for {symbol} between 9:30-9:45 AM")
            else:
                print(f"No data available for {symbol}")
        except Exception as e:
            print(f"Error fetching data for {symbol}: {e}")
    
    return data

def create_excel(stock_data):
    """Create an Excel file with the stock data and conditional formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Prices"
    
    # Define fill colors
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")    # Red
    
    # Add headers
    headers = ['Stock', '9:30-9:45 High', '9:30-9:45 Low', 'CMP', 'Status']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    # Add data
    row_num = 2
    for symbol, data in stock_data.items():
        ws.cell(row=row_num, column=1).value = symbol
        ws.cell(row=row_num, column=2).value = data['High']
        ws.cell(row=row_num, column=3).value = data['Low']
        ws.cell(row=row_num, column=4).value = data['CMP']
        
        # Calculate proximity to high and low
        high_proximity = abs((data['CMP'] - data['High']) / data['High'] * 100)
        low_proximity = abs((data['CMP'] - data['Low']) / data['Low'] * 100)
        
        # Apply conditional formatting
        if high_proximity <= THRESHOLD_PERCENTAGE:
            ws.cell(row=row_num, column=4).fill = green_fill
            ws.cell(row=row_num, column=5).value = "Near High"
        elif low_proximity <= THRESHOLD_PERCENTAGE:
            ws.cell(row=row_num, column=4).fill = red_fill
            ws.cell(row=row_num, column=5).value = "Near Low"
        else:
            ws.cell(row=row_num, column=5).value = "Normal"
        
        row_num += 1
    
    # Save the workbook
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{timestamp}_{OUTPUT_FILE}"
    wb.save(filename)
    print(f"Excel file created: {filename}")
    return filename

def run_tracker(stock_list=None):
    """Run the stock tracker and generate the Excel file."""
    print(f"Running stock tracker at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if stock_list is None:
        stock_list = STOCKS
    
    stock_data = get_stock_data(stock_list)
    if stock_data:
        filename = create_excel(stock_data)
        return filename
    else:
        print("No data was retrieved. Excel file not created.")
        return None

def schedule_tracker():
    """Schedule the tracker to run at 9:45 AM on weekdays."""
    schedule.every().monday.at("09:45").do(run_tracker)
    schedule.every().tuesday.at("09:45").do(run_tracker)
    schedule.every().wednesday.at("09:45").do(run_tracker)
    schedule.every().thursday.at("09:45").do(run_tracker)
    schedule.every().friday.at("09:45").do(run_tracker)
    
    print("Stock tracker scheduled to run at 9:45 AM on weekdays.")
    print("Press Ctrl+C to exit.")
    
    try:
        while True:
            schedule.run_pending()
            time.sleep(1)
    except KeyboardInterrupt:
        print("Scheduler stopped.")

def check_market_hours():
    """Check if the current time is within market hours."""
    now = datetime.datetime.now()
    # Market hours: Monday-Friday, 9:30 AM - 4:00 PM Eastern Time
    is_weekday = 0 <= now.weekday() <= 4
    is_market_hours = datetime.time(9, 30) <= now.time() <= datetime.time(16, 0)
    return is_weekday and is_market_hours

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Stock Tracker App using 9:30-9:45 AM candle")
    parser.add_argument("--stocks", nargs="+", help="List of stock symbols to track")
    parser.add_argument("--schedule", action="store_true", help="Schedule the tracker to run at 9:45 AM on weekdays")
    parser.add_argument("--run", action="store_true", help="Run the tracker immediately")
    parser.add_argument("--force", action="store_true", help="Force run even outside market hours")
    
    args = parser.parse_args()
    
    if args.schedule:
        schedule_tracker()
    elif args.run or not (args.schedule or args.stocks):
        # Check market hours unless --force is specified
        if args.force or check_market_hours():
            run_tracker(args.stocks)
        else:
            print("Market is currently closed. Use --force to run anyway.")
    else:
        # If only stocks are provided but no action, run immediately with those stocks
        if args.force or check_market_hours():
            run_tracker(args.stocks)
        else:
            print("Market is currently closed. Use --force to run anyway.")
